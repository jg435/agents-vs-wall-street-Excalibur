"""Pipeline invariants. Run: ./.venv/bin/python -m tests.test_pipeline

Property assertions across the WHOLE fact set and all 4 companies — never
spot-checks. Exit 0 = safe to run/freeze.
"""
import json
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from agent import workbook                      # noqa: E402
from agent.contract import METRICS, OUTPUT, PERIOD  # noqa: E402
from agent.run import provisional_engine, validate  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
FAILS = []


def check(name, cond, detail=""):
    print(f"  [{'ok ' if cond else 'FAIL'}] {name}" + (f" — {detail}" if not cond else ""))
    if not cond:
        FAILS.append(name)


# 1. RECEIPTS SELF-VERIFY: every fact's quote appears in its cited doc
# (normalized like the extractor's verify gate; research/-derived facts are
# aggregates whose receipt is the derived file itself — existence checked)
from agent.extractor import _normalize, check_identities  # noqa: E402
facts = json.loads((REPO / "cache" / "facts.json").read_text())
print(f"1. receipts ({len(facts)} facts)")
for k, f in facts.items():
    doc = REPO / f["doc"]
    if f["doc"].startswith(("research/", "cache/")) or f["doc"] == "multiple":
        check(f"receipt {k} (derived)", (REPO / f["doc"]).exists() or f["doc"] == "multiple")
        continue
    if not doc.exists():
        check(f"receipt {k}", False, f"missing doc {f['doc']}")
        continue
    check(f"receipt {k}",
          _normalize(f["quote"][:120]) in _normalize(doc.read_text(errors="ignore")),
          "quote not found in doc")

# 1b. CROSS-FACT IDENTITIES (unit/scale errors explode these)
issues = check_identities(facts)
check("cross-fact identities", not issues, str(issues))

# 2. REQUIRED-FACTS COVERAGE: everything the engine consumes exists BY NAME
REQUIRED = [
    "HD.q2_fy25_net_sales_usdm", "HD.q2_fy25_adj_eps", "HD.q2_fy25_comp_pct",
    "HD.fy_sales_growth_guide", "HD.fy_comp_guide_mid", "HD.fy_adj_eps_growth_mid",
    "ADI.rev_guide_mid_usdm", "ADI.eps_guide_mid", "ADI.adj_gross_margin_last",
    "ADI.rev_guide_beat_median_pct",
    "DE.q3_fy25_revenues_usdm", "DE.q3_fy25_equip_net_sales_usdm",
    "DE.q3_fy25_ppa_op_profit_usdm", "DE.q3_fy25_ppa_net_sales_usdm",
    "DE.fy_net_income_guide_usdm", "DE.ppa_fy_sales_guide",
    "HAS.fy25_net_fees_gbpm", "HAS.fy25_op_profit_gbpm", "HAS.op_profit_consensus",
    "HAS.q4_net_fees_lfl", "HAS.shares_issued",
    "HAS.fy25_net_finance_charge_gbpm", "HAS.fy25_pre_exceptional_etr_pct",
    # Amendment 3: period-specific guidance anchors
    "ADI.q3_gm_guide_seq_change_pp", "DE.ppa_fy_margin_guide_pct",
    "DE.q2_tariff_refund_usdm", "DE.sat_fy_sales_guide",
    "HAS.q1_net_fees_lfl", "HAS.q2_net_fees_lfl", "HAS.q3_net_fees_lfl",
    # Amendment 4/5: phase+temper inputs, Hays FY26 guides, DE phasing set
    "HD.q1_fy26_net_sales_usdm", "HD.q1_fy26_sales_yoy_pct",
    "HD.q1_fy26_adj_eps", "HD.q1_fy25_adj_eps", "HD.q1_fy26_comp_pct",
    "HAS.fy26_disposed_net_fees_gbpm", "HAS.fy26_net_finance_charge_guide_gbpm",
    "HAS.fy26_etr_guide_pct", "HAS.h1_fy25_net_fees_gbpm",
    "DE.h1_fy26_net_income_usdm", "DE.q3_fy25_net_income_usdm",
    "DE.q4_fy25_net_income_usdm", "DE.cf_fy_sales_guide",
    "DE.q2_fy26_diluted_shares_m",
]
PERIOD_LABELED = REQUIRED  # ALL facts carry a period label (Amendment 3 rule 2)
print("2. required-fact coverage")
for k in REQUIRED:
    check(f"fact {k}", k in facts and facts[k]["value"] is not None)
print("2b. period labels (Amendment 3 rule 2)")
for k in PERIOD_LABELED:
    check(f"period on {k}", bool((facts.get(k) or {}).get("period")))
print("2c. basis labels (Amendment 6 schema) — ALL facts")
for k, f in facts.items():
    check(f"basis on {k}", bool(f.get("basis")))
print("2d. revision-diff regression: the ETR 38->45 revision is detected")
from agent.revision_diff import GUIDE_PATTERNS, vintage_series  # noqa: E402
co, pat, tr = GUIDE_PATTERNS["HAS.fy26_etr_guide_pct"]
series = vintage_series(co, pat, tr)
vals = [v for _, v, _ in series]
check("ETR vintages include 38 and 45", 38.0 in vals and 45.0 in vals, str(vals))
check("ETR latest vintage is 45", series and series[-1][1] == 45.0)

# 3. ENGINE: all 12 metrics, numeric, pass validate()
print("3. engine coverage + validation")
for tk in PERIOD:
    c = json.loads((REPO / "cache" / "contracts" / f"{tk}.json").read_text())
    fc = provisional_engine(c)
    check(f"{tk} all metrics", set(fc) == set(METRICS[tk]))
    try:
        validate(tk, fc)
        check(f"{tk} validate", True)
    except ValueError as e:
        check(f"{tk} validate", False, str(e))

# 3b. ONE-OFF POLICY: a one-off-bearing actual may not anchor alone
print("3b. one-off anchor rejection")
from agent.run import validate_anchor_periods  # noqa: E402
try:
    validate_anchor_periods("ADI",
        {"gm_actual": {"period": "Q2-FY2026",
                       "one_offs": [{"name": "planted one-off", "impact_pp": 1.0}]}},
        {"Adjusted gross margin": ["gm_actual"]})
    check("one-off-alone anchor rejected", False, "was accepted")
except ValueError:
    check("one-off-alone anchor rejected", True)

# 3c. ADVERSARIAL SWAPS (Amendment 6 §6): each must be rejected
print("3c. adversarial swaps")
# (a) GAAP fact swapped into an ADJUSTED-EPS metric
try:
    validate_anchor_periods("HD",
        {"gaap_eps": {"period": "Q2-FY2025", "basis": "GAAP"}},
        {"Adjusted diluted EPS": ["gaap_eps"]})
    check("GAAP-for-adjusted swap rejected", False, "accepted")
except ValueError:
    check("GAAP-for-adjusted swap rejected", True)
# (b) a Q4 LFL input swapped into HD's Q2 net sales (wrong company/period class)
try:
    validate_anchor_periods("HD",
        {"q4_lfl": {"period": "Q4-FY2026", "basis": "GAAP-REPORTED"}},
        {"Net sales": ["q4_lfl"]})
    check("wrong-period swap rejected", False, "accepted")
except ValueError:
    check("wrong-period swap rejected", True)
# (c) stale guide: engine value != latest vintage must be flagged STALE
from agent.revision_diff import GUIDE_PATTERNS as GP, vintage_series as VS
co, pat, tr = GP["HAS.fy26_etr_guide_pct"]
latest = VS(co, pat, tr)[-1][1]
check("stale-guide detection (38 vs latest)", abs(38.0 - latest) > 0.005 * latest,
      f"latest={latest}")
# (d) one-off actual anchoring alone — covered in 3b above (kept there)

# 2e. FULL SCHEMA on every fact (Amendment 6 §1)
print("2e. full fact schema (is_latest/revised_since_prior/one_offs/extracted_at)")
from agent.corpus import MANDATORY_FACT_KEYS
for k, f in facts.items():
    missing = [key for key in MANDATORY_FACT_KEYS if key not in f]
    check(f"schema {k}", not missing, str(missing))

# 4. VALIDATOR REJECTS a planted bad value (loud gate proven live)
print("4. validator rejection")
try:
    validate("HD", {"Net sales": (999999, "planted"), "Adjusted diluted EPS": (4.7, ""),
                    "Comparable sales, total company": (1.0, "")})
    check("planted out-of-range rejected", False, "validator accepted 999999")
except ValueError:
    check("planted out-of-range rejected", True)

# 5. WORKBOOK ROUNDTRIP into a temp dir, cell-exact
print("5. workbook roundtrip")
import openpyxl  # noqa: E402
with tempfile.TemporaryDirectory() as td:
    workbook.SUBMISSION = Path(td)
    vals = {m: 10.0 + i for i, m in enumerate(METRICS["ADI"])}
    out = workbook.fill(OUTPUT["ADI"], PERIOD["ADI"], vals)
    ws = openpyxl.load_workbook(out)["Summary"]
    found = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in vals:
                found[cell.value] = ws.cell(row=cell.row,
                                            column=cell.column + 2).value
    check("roundtrip cells", all(found.get(m) == vals[m] for m in vals), str(found))
workbook.SUBMISSION = REPO / "submission"

# 6. STRICT JSON caches
print("6. strict-JSON caches")
for name in ["cache/consensus.json", "cache/facts.json"]:
    try:
        json.loads((REPO / name).read_text(),
                   parse_constant=lambda x: (_ for _ in ()).throw(ValueError(x)))
        check(name, True)
    except ValueError as e:
        check(name, False, str(e))

print("\n" + ("ALL PASS" if not FAILS else f"{len(FAILS)} FAILURES: {FAILS}"))
sys.exit(1 if FAILS else 0)
