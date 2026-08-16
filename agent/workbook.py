"""xlsx writer — fills ONLY the forecast cells of the supplied templates.

Contract (SUBMISSION.md): keep the Summary sheet, metric labels, units and
fiscal-period header untouched. Percentages in percentage points; Hays EPS
in pence. Loud failure on any label/period mismatch — never silent.
"""
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
TEMPLATES = REPO / "challenge" / "templates"
SUBMISSION = REPO / "submission"


def fill(output_file, period, forecasts):
    """forecasts: {metric_label: numeric value}. Finds each metric row by
    exact label and the forecast column by the period header."""
    wb = openpyxl.load_workbook(TEMPLATES / output_file)
    ws = wb["Summary"]

    period_col, metric_rows = None, {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == period:
                period_col = cell.column
            if cell.value in forecasts:
                metric_rows[cell.value] = cell.row

    if period_col is None:
        raise ValueError(f"{output_file}: period header '{period}' not found")
    missing = set(forecasts) - set(metric_rows)
    if missing:
        raise ValueError(f"{output_file}: metric labels not found: {missing}")

    for label, value in forecasts.items():
        if not isinstance(value, (int, float)) or value != value:  # NaN guard
            raise ValueError(f"{output_file}: non-numeric forecast for '{label}': {value!r}")
        ws.cell(row=metric_rows[label], column=period_col, value=round(float(value), 2))

    SUBMISSION.mkdir(exist_ok=True)
    out = SUBMISSION / output_file
    wb.save(out)
    return out
